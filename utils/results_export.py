"""
Results Export Module - Simplified Event Log Export

Handles hourly segment export of event logs only:
- Event log data per hourly segment
- Multiple output formats (JSON, CSV, Excel)
- Master event log that appends all events to single Excel file and JSON log
- Optional PostgreSQL cloud upload
- No summaries or aggregations

Uses a queue-based writer for non-blocking master log updates.
"""

import json
import csv
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass
import logging
import datetime
import threading
import time
import copy
from dataclasses import asdict
from queue import Queue, Empty
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from configparser import ConfigParser
import requests
import os
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor
import math
import uuid



# Global lock for master log file access (shared across all ResultsExporter instances)
_master_log_lock = threading.Lock()

# ============================================================================
# QUEUE-BASED MASTER LOG WRITER
# Workers put events in a queue and continue immediately (non-blocking).
# A single background thread handles all file I/O.
# ============================================================================

class MasterLogWriter:
    """
    Singleton background writer for master log.
    Collects events from all workers via queue and writes in batches.
    Workers never block waiting for file I/O.
    """
    _instance = None
    _lock = threading.Lock()

    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return

        self._initialized = True
        self.logger = logging.getLogger(__name__ + ".writer")

        # Queue for pending writes
        self._write_queue = Queue()

        # Queue for dwell time updates
        self._dwell_queue = Queue()

        # Writer thread
        self._writer_thread = None
        self._stop_event = threading.Event()

        # Batch settings
        self._batch_interval = 300.0  # Write batches every 5 minutes
        self._max_batch_size = 500  # Or when batch reaches this size

        # Track master log paths (set by first exporter)
        self._master_log_path = None
        self._master_json_path = None

    def set_master_log_path(self, path: Path, json_path: Path = None):
        """Set the master log paths (called by ResultsExporter)"""
        if self._master_log_path is None:
            self._master_log_path = path
            self._master_json_path = json_path
            self._start_writer()

    def _start_writer(self):
        """Start the background writer thread"""
        if self._writer_thread is None or not self._writer_thread.is_alive():
            self._stop_event.clear()
            self._writer_thread = threading.Thread(
                target=self._writer_loop,
                daemon=True,
                name="MasterLogWriter"
            )
            self._writer_thread.start()
            self.logger.info("Master log writer thread started")

    def queue_events(self, events: List[Dict], video_source: str, segment_id: str):
        """
        Queue events for writing (NON-BLOCKING).
        Workers call this and continue immediately.
        """
        if events:
            self._write_queue.put({
                'events': events,
                'video_source': video_source,
                'segment_id': segment_id,
                'timestamp': time.time()
            })

    def queue_dwell_update(self, updates: List[Dict], video_source: str):
        """Queue dwell time updates (NON-BLOCKING)"""
        if updates:
            self._dwell_queue.put({
                'updates': updates,
                'video_source': video_source,
                'timestamp': time.time()
            })

    def _writer_loop(self):
        """Background thread that processes queued writes"""
        self.logger.info("Writer loop started")

        while not self._stop_event.is_set():
            try:
                # Collect batch of events
                batch_events = []
                batch_start = time.time()

                # Collect events until batch interval or max size
                while (time.time() - batch_start < self._batch_interval and
                       len(batch_events) < self._max_batch_size):
                    try:
                        item = self._write_queue.get(timeout=0.5)
                        batch_events.append(item)
                    except Empty:
                        pass

                    if self._stop_event.is_set():
                        break

                # Write batch if we have events
                if batch_events:
                    self._write_batch(batch_events)

                # Process dwell updates (less frequently)
                self._process_dwell_updates()

            except Exception as e:
                self.logger.error(f"Writer loop error: {e}")
                time.sleep(1)

        # Flush remaining on shutdown
        self._flush_all()
        self.logger.info("Writer loop stopped")

    def _write_batch(self, batch_items: List[Dict]):
        """Write a batch of events to both the Excel and JSON master logs"""
        if not self._master_log_path or not batch_items:
            return

        # Prepare all rows and JSON entries
        all_rows = []
        json_entries = []
        for item in batch_items:
            for event in item['events']:
                event_type = 'line_crossing' if 'line_name' in event else 'zone_entry' if 'zone_name' in event else 'unknown'
                dwell_val = event.get('dwell_seconds', 0.0)
                if dwell_val is not None:
                    try:
                        dwell_val = float(dwell_val)
                        if dwell_val < 0.3 or dwell_val > 86400:
                            dwell_val = ''
                    except:
                        dwell_val = ''
                else:
                    dwell_val = ''

                row = (
                    event.get('actual_datetime', ''),
                    event_type,
                    event.get('track_id', ''),
                    event.get('class_id', ''),
                    event.get('class_name', ''),
                    event.get('line_name', ''),
                    event.get('zone_name', ''),
                    event.get('direction', ''),
                    event.get('confidence', ''),
                    event.get('speed', 0.0),
                    event.get('speed_units', ''),
                    dwell_val,
                    item['video_source'] or '',
                    item['segment_id'] if item['segment_id'] is not None else ''
                )
                all_rows.append(row)

                # Also prepare JSON entry
                json_entries.append({
                    'actual_datetime': event.get('actual_datetime', ''),
                    'event_type': event_type,
                    'track_id': event.get('track_id', ''),
                    'class_id': event.get('class_id', ''),
                    'class_name': event.get('class_name', ''),
                    'line_name': event.get('line_name', ''),
                    'zone_name': event.get('zone_name', ''),
                    'direction': event.get('direction', ''),
                    'confidence': event.get('confidence', ''),
                    'speed': event.get('speed', 0.0),
                    'speed_units': event.get('speed_units', ''),
                    'dwell_seconds': dwell_val if dwell_val != '' else None,
                    'video_source': item['video_source'] or '',
                    'segment_id': item['segment_id'] if item['segment_id'] is not None else ''
                })

        if not all_rows:
            return

        # Write to files with lock
        max_retries = 3
        for attempt in range(max_retries):
            if not _master_log_lock.acquire(timeout=10):
                self.logger.warning(f"Could not acquire lock for batch write (attempt {attempt + 1})")
                continue

            try:
                # Write to Excel
                if not self._master_log_path.exists():
                    # Create file with headers
                    df = pd.DataFrame(columns=[
                        'actual_datetime', 'event_type', 'track_id',
                        'class_id', 'class_name', 'line_name', 'zone_name',
                        'direction', 'confidence', 'speed', 'speed_units', 'dwell_seconds',
                        'video_source', 'segment_id'
                    ])
                    df.to_excel(self._master_log_path, sheet_name='Events', index=False, engine='openpyxl')

                wb = load_workbook(self._master_log_path)
                ws = wb.active
                current_rows = ws.max_row

                for row in all_rows:
                    ws.append(row)

                wb.save(self._master_log_path)
                wb.close()

                # Write to JSON (append mode, newline-delimited JSON)
                if self._master_json_path:
                    try:
                        with open(self._master_json_path, 'a') as f:
                            for entry in json_entries:
                                f.write(json.dumps(entry, default=str) + "\n")
                    except Exception as json_err:
                        self.logger.warning(f"Failed to write to JSON master log: {json_err}")

                self.logger.info(f"Batch wrote {len(all_rows)} events to master logs (total: {current_rows + len(all_rows) - 1})")
                _master_log_lock.release()
                return

            except PermissionError:
                self.logger.warning("Permission error, retrying...")
                _master_log_lock.release()
                time.sleep(0.5)
            except Exception as e:
                self.logger.error(f"Batch write failed: {e}")
                _master_log_lock.release()
                return

        self.logger.error("Failed to write batch after retries")

    def _process_dwell_updates(self):
        """Process queued dwell time updates"""
        updates_by_source = {}

        # Collect all pending updates
        while True:
            try:
                item = self._dwell_queue.get_nowait()
                source = item['video_source']
                if source not in updates_by_source:
                    updates_by_source[source] = []
                updates_by_source[source].extend(item['updates'])
            except Empty:
                break

        if not updates_by_source:
            return

        # Apply updates (this still needs read-modify-write unfortunately)
        if not _master_log_lock.acquire(timeout=5):
            # Re-queue for next cycle
            for source, updates in updates_by_source.items():
                self._dwell_queue.put({'updates': updates, 'video_source': source, 'timestamp': time.time()})
            return

        try:
            if not self._master_log_path or not self._master_log_path.exists():
                return

            df = pd.read_excel(self._master_log_path, sheet_name='Events', engine='openpyxl')
            if df.empty:
                return

            updates_made = 0
            for source, updates in updates_by_source.items():
                for update in updates:
                    track_id = update.get('track_id')
                    zone_name = update.get('zone_name')
                    new_dwell = update.get('dwell_seconds', 0.0)

                    if not track_id or not zone_name:
                        continue

                    try:
                        dwell = float(new_dwell) if new_dwell is not None else 0.0
                        if dwell < 0 or dwell > 86400:
                            continue
                    except:
                        continue

                    mask = (
                        (df['track_id'] == track_id) &
                        (df['zone_name'] == zone_name) &
                        (df['event_type'] == 'zone_entry')
                    )
                    if source:
                        mask = mask & (df['video_source'] == source)

                    if mask.any():
                        df.loc[mask, 'dwell_seconds'] = dwell
                        updates_made += 1

            if updates_made > 0:
                with pd.ExcelWriter(self._master_log_path, engine='openpyxl', mode='w') as writer:
                    df.to_excel(writer, sheet_name='Events', index=False)
                self.logger.debug(f"Updated dwell times for {updates_made} events")

        except Exception as e:
            self.logger.error(f"Dwell update failed: {e}")
        finally:
            _master_log_lock.release()

    def _flush_all(self):
        """Flush all remaining events on shutdown"""
        # Collect all remaining events
        remaining = []
        while True:
            try:
                remaining.append(self._write_queue.get_nowait())
            except Empty:
                break

        if remaining:
            self._write_batch(remaining)

        # Process remaining dwell updates
        self._process_dwell_updates()

    def flush_and_wait(self, timeout: float = 10.0):
        """Force flush and wait for completion (call at end of processing)"""
        # Signal we want to flush
        flush_complete = threading.Event()

        def flush_task():
            self._flush_all()
            flush_complete.set()

        threading.Thread(target=flush_task, daemon=True).start()
        flush_complete.wait(timeout=timeout)

    def stop(self):
        """Stop the writer thread"""
        self._stop_event.set()
        if self._writer_thread and self._writer_thread.is_alive():
            self._writer_thread.join(timeout=15)


# Global writer instance
_master_log_writer = None

def get_master_log_writer() -> MasterLogWriter:
    """Get the global master log writer instance"""
    global _master_log_writer
    if _master_log_writer is None:
        _master_log_writer = MasterLogWriter()
    return _master_log_writer


@dataclass
class ApiUploadBatch:
    """A sealed group of events that must be uploaded exactly once by the client."""
    batch_id: str
    events: List[Dict]
    created_at: float
    reason: str


class ApiBatchUploader:
    """
    Process-wide API uploader aligned to wall-clock five-minute boundaries.

    Producers append events to the current pending buffer without waiting for
    network I/O. At each boundary (for example 13:00, 13:05, 13:10), the
    pending list is atomically swapped into a sealed batch. New events then go
    into a fresh list while the sealed batch uploads on a separate thread.
    """

    def __init__(self, interval_seconds: int = 300, max_retry_batches: int = 24):
        self.logger = logging.getLogger(__name__ + ".api_uploader")
        self.interval_seconds = interval_seconds
        self.max_retry_batches = max(1, int(max_retry_batches))

        self._api_url = None
        self._api_key = None
        self._output_folder = None

        self._pending_events = []
        self._retry_batches = []
        self._pending_lock = threading.Lock()
        self._upload_queue = Queue()

        self._stop_event = threading.Event()
        self._scheduler_thread = None
        self._uploader_thread = None
        self._start_lock = threading.Lock()
        self._rejection_lock = threading.Lock()

        self._started = False
        self._finalizing = False
        self._shutdown_complete = False

    @property
    def shutdown_complete(self) -> bool:
        return self._shutdown_complete

    @staticmethod
    def _has_value(value) -> bool:
        return value is not None and str(value).strip() != ''

    @classmethod
    def _infer_event_type(cls, event: Dict) -> str:
        event_type = event.get('event_type')
        if cls._has_value(event_type):
            return str(event_type)
        if cls._has_value(event.get('line_name')):
            return 'line_crossing'
        if cls._has_value(event.get('zone_name')):
            return 'zone_entry'
        return 'unknown'

    def configure(self, api_url: str, api_key: str, output_folder: Path):
        """Configure and start the shared uploader. Safe to call repeatedly."""
        if not api_url or not api_key:
            raise ValueError("API_URL and API_KEY are required for API uploads")

        with self._start_lock:
            if self._started:
                if api_url != self._api_url:
                    self.logger.warning("API uploader is already running with a different API_URL; keeping the original URL")
                return

            self._api_url = api_url
            self._api_key = api_key
            self._output_folder = Path(output_folder)
            self._output_folder.mkdir(parents=True, exist_ok=True)

            self._stop_event.clear()
            self._finalizing = False
            self._shutdown_complete = False

            self._scheduler_thread = threading.Thread(
                target=self._scheduler_loop,
                daemon=True,
                name="ApiUploadScheduler"
            )
            self._uploader_thread = threading.Thread(
                target=self._uploader_loop,
                daemon=True,
                name="ApiUploadWorker"
            )

            self._uploader_thread.start()
            self._scheduler_thread.start()
            self._started = True

            next_boundary = self._next_boundary_epoch()
            next_label = datetime.datetime.fromtimestamp(next_boundary).strftime('%Y-%m-%d %H:%M:%S')
            self.logger.info(
                f"API batch uploader started; uploads align to five-minute boundaries. "
                f"Next boundary: {next_label}"
            )

    def queue_events(self, events: List[Dict]) -> int:
        """Append events to the active interval buffer without blocking on uploads."""
        if not events or not self._started:
            return 0

        events_copy = copy.deepcopy(events)
        with self._pending_lock:
            if self._finalizing:
                self.logger.warning("API uploader is shutting down; ignored events queued after final flush began")
                return 0
            self._pending_events.extend(events_copy)
            return len(events_copy)

    def _next_boundary_epoch(self, now: float = None) -> float:
        now = time.time() if now is None else now
        return (math.floor(now / self.interval_seconds) + 1) * self.interval_seconds

    def _scheduler_loop(self):
        while not self._stop_event.is_set():
            boundary_epoch = self._next_boundary_epoch()
            wait_seconds = max(0.0, boundary_epoch - time.time())

            if self._stop_event.wait(timeout=wait_seconds):
                break

            boundary_label = datetime.datetime.fromtimestamp(boundary_epoch).strftime('%Y-%m-%d %H:%M:%S')
            self._seal_pending_batch(reason=f"scheduled boundary {boundary_label}")

    def _seal_pending_batch(self, reason: str) -> int:
        """Atomically detach pending events and enqueue immutable upload batches."""
        with self._pending_lock:
            retry_batches = self._retry_batches
            self._retry_batches = []

            events = self._pending_events
            self._pending_events = []

        total_events = 0

        # Retry sealed batches without changing their batch IDs. This lets a
        # server that supports X-Idempotency-Key recognize a retry as the same
        # logical upload rather than a brand-new batch.
        for retry_batch in retry_batches:
            self._upload_queue.put(retry_batch)
            total_events += len(retry_batch.events)
            self.logger.info(
                f"API batch {retry_batch.batch_id[:8]} queued for retry ({reason})."
            )

        if events:
            batch = ApiUploadBatch(
                batch_id=uuid.uuid4().hex,
                events=events,
                created_at=time.time(),
                reason=reason
            )
            self._upload_queue.put(batch)
            total_events += len(events)
            self.logger.info(
                f"API batch {batch.batch_id[:8]} sealed with {len(events)} events ({reason})."
            )

        return total_events

    def _uploader_loop(self):
        while True:
            item = self._upload_queue.get()
            try:
                if item is None:
                    return

                success = self._upload_batch(item)
                if not success:
                    if self._finalizing:
                        self._persist_failed_batch(item, "final upload failed")
                    else:
                        self._retain_retry_batch(item)
                        self.logger.warning(
                            f"API batch {item.batch_id[:8]} will retry at the next "
                            f"upload boundary."
                        )
            except Exception as exc:
                self.logger.exception(f"Unexpected API uploader error: {exc}")
                if item is not None:
                    self._persist_failed_batch(item, f"unexpected uploader error: {exc}")
            finally:
                self._upload_queue.task_done()

    def _retain_retry_batch(self, batch: ApiUploadBatch) -> None:
        """Keep a bounded retry backlog and persist overflow instead of leaking memory."""
        overflow_batches = []
        with self._pending_lock:
            self._retry_batches.append(batch)
            while len(self._retry_batches) > self.max_retry_batches:
                overflow_batches.append(self._retry_batches.pop(0))
        for overflow_batch in overflow_batches:
            self._persist_failed_batch(
                overflow_batch,
                "retry backlog exceeded in-memory limit",
            )

    def _prepare_payload(self, event_list: List[Dict]):
        """Return (valid_payload, rejected_events). Invalid track IDs are excluded."""
        valid_events = []
        rejected_events = []

        for event in event_list:
            normalized = copy.deepcopy(event)
            normalized["event_type"] = self._infer_event_type(normalized)
            if not self._has_value(normalized.get("video_source")):
                normalized["video_source"] = "unknown"

            raw_track_id = normalized.get("track_id")

            try:
                if raw_track_id is None or isinstance(raw_track_id, bool):
                    raise ValueError("track_id is missing")
                numeric_track_id = float(raw_track_id)
                if not math.isfinite(numeric_track_id):
                    raise ValueError("track_id is not finite")
                if not numeric_track_id.is_integer():
                    raise ValueError("track_id is not an integer")
                normalized["track_id"] = int(numeric_track_id)
                valid_events.append(normalized)
            except (TypeError, ValueError):
                rejected_events.append({
                    "reason": "invalid or missing track_id",
                    "event": normalized
                })

        if not valid_events:
            return [], rejected_events

        df = pd.DataFrame(valid_events)

        drop_cols = [
            "event_id", "class_id", "speed_units", "segment_id",
            "confidence", "timestamp", "position"
        ]
        df = df.drop(columns=[c for c in drop_cols if c in df.columns])
        df = df.rename(columns={"actual_datetime": "time", "speed": "speed_mph"})

        if "time" in df.columns:
            df["time"] = df["time"].apply(
                lambda value: value.isoformat() if hasattr(value, "isoformat") else value
            )

        if "speed_mph" in df.columns:
            df["speed_mph"] = pd.to_numeric(df["speed_mph"], errors="coerce").round().astype("Int64")

        if "dwell_seconds" in df.columns:
            dwell = pd.to_numeric(df["dwell_seconds"], errors="coerce").fillna(0)
            df["dwell_seconds"] = np.ceil(dwell).astype(int)

        # Convert pandas NA/NaN/Infinity values to plain None before requests
        # serializes the payload.
        df = df.replace([np.inf, -np.inf], np.nan)
        df = df.astype(object).where(pd.notna(df), None)

        return df.to_dict(orient="records"), rejected_events

    def _upload_batch(self, batch: ApiUploadBatch) -> bool:
        payload, rejected_events = self._prepare_payload(batch.events)

        if rejected_events:
            self._persist_rejected_events(rejected_events, batch.batch_id)
            self.logger.warning(
                f"API batch {batch.batch_id[:8]} excluded {len(rejected_events)} event(s) "
                f"with invalid track_id values. Valid events will still upload."
            )

        if not payload:
            self.logger.warning(
                f"API batch {batch.batch_id[:8]} contained no valid records to upload."
            )
            return True

        headers = {
            "Content-Type": "application/json",
            "X-API-Key": self._api_key,
            # The server may use this header for idempotency now or later. The
            # same key is retained for this sealed batch.
            "X-Idempotency-Key": batch.batch_id
        }

        try:
            safe_payload = self._sanitize_for_json(payload)
            response = requests.post(
                self._api_url,
                json=safe_payload,
                headers=headers,
                timeout=30
            )

            if 200 <= response.status_code < 300:
                self.logger.info(
                    f"Cloud API: Successfully uploaded {len(payload)} records "
                    f"from batch {batch.batch_id[:8]}."
                )
                return True

            self.logger.error(
                f"Cloud API Error for batch {batch.batch_id[:8]} "
                f"({response.status_code}): {response.text}"
            )
            return False

        except requests.exceptions.RequestException as exc:
            self.logger.error(
                f"Cloud API Connection Failed for batch {batch.batch_id[:8]}: {exc}"
            )
            return False

    def _sanitize_for_json(self, data):
        if isinstance(data, dict):
            return {key: self._sanitize_for_json(value) for key, value in data.items()}
        if isinstance(data, list):
            return [self._sanitize_for_json(value) for value in data]
        if isinstance(data, float) and (math.isnan(data) or math.isinf(data)):
            return None
        return data

    def _persist_rejected_events(self, rejected_events: List[Dict], batch_id: str):
        if not self._output_folder or not rejected_events:
            return

        path = self._output_folder / "api_rejected_events.jsonl"
        timestamp = datetime.datetime.now().isoformat()
        with self._rejection_lock:
            with open(path, "a", encoding="utf-8") as handle:
                for rejected in rejected_events:
                    handle.write(json.dumps({
                        "rejected_at": timestamp,
                        "batch_id": batch_id,
                        **rejected
                    }, default=str) + "\n")

    def _persist_failed_batch(self, batch: ApiUploadBatch, reason: str):
        if not self._output_folder:
            return

        path = self._output_folder / "api_failed_batches.jsonl"
        with self._rejection_lock:
            with open(path, "a", encoding="utf-8") as handle:
                handle.write(json.dumps({
                    "failed_at": datetime.datetime.now().isoformat(),
                    "batch_id": batch.batch_id,
                    "reason": reason,
                    "event_count": len(batch.events),
                    "events": batch.events
                }, default=str) + "\n")
        self.logger.error(
            f"Saved failed API batch {batch.batch_id[:8]} to {path}."
        )

    def shutdown(self):
        """Seal remaining events, upload them, and wait for all API work to finish."""
        with self._start_lock:
            if not self._started or self._shutdown_complete:
                return
            self._finalizing = True
            self._stop_event.set()

        if self._scheduler_thread and self._scheduler_thread.is_alive():
            self._scheduler_thread.join(timeout=5)

        remaining = self._seal_pending_batch(reason="program shutdown final flush")
        if remaining:
            self.logger.info(f"Final API flush queued {remaining} leftover event(s).")
        else:
            self.logger.info("Final API flush found no leftover events.")

        # The sentinel is placed after all sealed batches, so the uploader exits
        # only after every scheduled and final batch has been attempted.
        self._upload_queue.put(None)
        self._upload_queue.join()

        if self._uploader_thread and self._uploader_thread.is_alive():
            self._uploader_thread.join(timeout=5)

        self._shutdown_complete = True
        self._started = False
        self.logger.info("API batch uploader stopped after final flush.")


_api_batch_uploader = None
_api_batch_uploader_lock = threading.Lock()


def get_api_batch_uploader() -> ApiBatchUploader:
    global _api_batch_uploader
    with _api_batch_uploader_lock:
        if _api_batch_uploader is None or _api_batch_uploader.shutdown_complete:
            _api_batch_uploader = ApiBatchUploader(interval_seconds=300)
        return _api_batch_uploader


@dataclass
class ExportConfig:
    export_formats: List[str] = None  # ['json', 'csv', 'excel']
    enable_master_log: bool = True  # Enable master event log
    enable_api_upload: bool = False

    def __post_init__(self):
        if self.export_formats is None:
            self.export_formats = ['json', 'csv', 'excel']


class ResultsExporter:
    """Handles export of event logs in multiple formats"""

    def __init__(self, output_folder: str, export_config: Optional[ExportConfig] = None):
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(parents=True, exist_ok=True)

        self.config = export_config or ExportConfig()
        self.logger = logging.getLogger(__name__)

        # Create segments folder for hourly exports
        self.segments_folder = self.output_folder / "segments"
        self.segments_folder.mkdir(exist_ok=True)

        # Master event log file
        self.master_log_path = self.output_folder / "master_event_log.xlsx"
        self.master_json_path = self.output_folder / "master_event_log.json"

        # NEW: Create a background worker specifically for file writing
        self.file_writer_pool = ThreadPoolExecutor(max_workers=2)

        self._file_writer_shutdown = False

        # Initialize master log if it doesn't exist
        if self.config.enable_master_log:
            self._initialize_master_log()
            writer = get_master_log_writer()
            writer.set_master_log_path(self.master_log_path, self.master_json_path)

        # Start one process-wide, wall-clock-aligned API uploader. Every worker
        # writes into the same pending interval buffer.
        if self.config.enable_api_upload:
            load_dotenv()
            api_url = os.getenv("API_URL")
            api_key = os.getenv("API_KEY")
            if api_url and api_key:
                get_api_batch_uploader().configure(
                    api_url=api_url,
                    api_key=api_key,
                    output_folder=self.output_folder
                )
            else:
                self.logger.error(
                    "API Upload is enabled, but API_URL or API_KEY is missing from .env"
                )

        self.logger.info(f"Results exporter initialized: {self.output_folder}")
        if self.config.enable_master_log:
            self.logger.info(f"Master event log: {self.master_log_path}")

    def shutdown(self, finalize_shared: bool = False):
        """
        Wait for this exporter's file tasks. The orchestrator passes
        finalize_shared=True once, after all workers have stopped, to flush the
        shared API buffer and master-log writer.
        """
        if not self._file_writer_shutdown:
            self.logger.info("Waiting for background file exports to finish saving...")
            self.file_writer_pool.shutdown(wait=True)
            self._file_writer_shutdown = True

        if finalize_shared:
            if self.config.enable_api_upload:
                try:
                    get_api_batch_uploader().shutdown()
                except Exception as exc:
                    self.logger.error(f"Final API flush failed: {exc}")

            if self.config.enable_master_log:
                try:
                    get_master_log_writer().stop()
                except Exception as exc:
                    self.logger.debug(f"Master log writer shutdown notice: {exc}")

        self.logger.info("Background exports safely finished.")

    def _sanitize_for_json(self, data):
        """Recursively hunts down NaN/Infinity values and converts them to None for safe JSON transport"""
        if isinstance(data, dict):
            return {k: self._sanitize_for_json(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._sanitize_for_json(v) for v in data]
        elif isinstance(data, float):
            if math.isnan(data) or math.isinf(data):
                return None
            return data
        return data

    def _initialize_master_log(self):
        """Initialize the master event log Excel file if it doesn't exist"""
        if not self.master_json_path.exists():
            # Create new file
            self.master_json_path.touch()
            self.logger.info(f"Created master JSON event log: {self.master_json_path}")

        if not self.master_log_path.exists():
            # Create new file with headers
            df = pd.DataFrame(columns=[
                'actual_datetime', 'event_type', 'track_id',
                'class_id', 'class_name', 'line_name', 'zone_name',
                'direction', 'confidence', 'speed', 'speed_units', 'dwell_seconds',
                'video_source', 'segment_id'
            ])
            df.to_excel(self.master_log_path, sheet_name='Events', index=False, engine='openpyxl')
            self.logger.info(f"Created master event log: {self.master_log_path}")

    def _append_to_master_log(self, event_list: List[Dict], video_source: str = None, segment_id: Union[int, str] = None):
        """
        Queue events for appending to master log (NON-BLOCKING).

        Events are queued and written in batches by a background thread.
        Both Excel and JSON master logs are updated at the same frequency.
        This allows workers to continue processing without waiting for file I/O.

        Args:
            event_list: List of event dictionaries to append
            video_source: Optional source video filename
            segment_id: Optional segment identifier
        """
        if not self.config.enable_master_log or not event_list:
            return

        # Queue events for background writing (non-blocking)
        events_copy = copy.deepcopy(event_list)

        writer = get_master_log_writer()
        writer.queue_events(
            events_copy,
            video_source or '',
            str(segment_id) if segment_id is not None else ''
        )

        self.logger.debug(f"Queued {len(event_list)} events for master log")

    def update_zone_dwell_times(self, zone_updates: List[Dict], video_source: str = None):
        """
        Queue dwell time updates for zone events (NON-BLOCKING).
        Updates are processed by the background writer thread.

        Args:
            zone_updates: List of dicts with 'track_id', 'zone_name', 'dwell_seconds'
            video_source: Video source to match (for uniqueness)
        """
        if not self.config.enable_master_log or not zone_updates:
            return

        # Queue updates for background processing (non-blocking)
        updates_copy = copy.deepcopy(zone_updates)

        writer = get_master_log_writer()
        writer.queue_dwell_update(updates_copy, video_source or '')

        self.logger.debug(f"Queued {len(zone_updates)} dwell updates for master log")

    def _sanitize_dwell_time(self, dwell_value, min_dwell: float = 0.3) -> Optional[float]:
        """
        Validate and sanitize dwell time values.
        Returns None for invalid values (negative, below minimum, or unreasonably large).

        Args:
            dwell_value: Raw dwell time value
            min_dwell: Minimum dwell time threshold (default 0.3 seconds)
        """
        try:
            dwell = float(dwell_value) if dwell_value is not None else 0.0
            # Invalid if negative or greater than 24 hours (86400 seconds)
            if dwell < 0 or dwell > 86400:
                return None
            # Filter below minimum threshold
            if dwell < min_dwell:
                return None
            return round(dwell, 2)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _has_value(value) -> bool:
        return value is not None and str(value).strip() != ''

    @classmethod
    def _infer_event_type(cls, event: Dict) -> str:
        event_type = event.get('event_type')
        if cls._has_value(event_type):
            return str(event_type)
        if cls._has_value(event.get('line_name')):
            return 'line_crossing'
        if cls._has_value(event.get('zone_name')):
            return 'zone_entry'
        return 'unknown'

    def _convert_events_to_dicts(
        self,
        events_list: List,
        video_source: str = None,
        segment_id: Union[int, str] = None
    ) -> List[Dict]:
        """Convert event objects to independent dictionaries for files/API queues."""
        event_dicts = []

        for evt in events_list:
            if hasattr(evt, 'to_dict'):
                data = evt.to_dict()
            elif hasattr(evt, '__dataclass_fields__'):
                data = asdict(evt)
            elif isinstance(evt, dict):
                data = copy.deepcopy(evt)
            else:
                data = {
                    'track_id': getattr(evt, 'track_id', None),
                    'class_id': getattr(evt, 'class_id', ''),
                    'class_name': getattr(evt, 'class_name', ''),
                    'line_name': getattr(evt, 'line_name', ''),
                    'zone_name': getattr(evt, 'zone_name', ''),
                    'direction': getattr(evt, 'direction', ''),
                    'confidence': getattr(evt, 'confidence', ''),
                    'speed': getattr(evt, 'avg_speed', 0.0),
                    'speed_units': getattr(evt, 'speed_units', ''),
                    'dwell_seconds': getattr(evt, 'dwell_seconds', 0.0),
                    'actual_datetime': getattr(evt, 'actual_datetime', '')
                }

            actual_datetime = data.get('actual_datetime')
            if actual_datetime and hasattr(actual_datetime, 'isoformat'):
                data['actual_datetime'] = actual_datetime.isoformat()

            data['dwell_seconds'] = self._sanitize_dwell_time(
                data.get('dwell_seconds', 0.0)
            )
            data['event_type'] = self._infer_event_type(data)
            if video_source is not None and not self._has_value(data.get('video_source')):
                data['video_source'] = video_source
            if segment_id is not None and not self._has_value(data.get('segment_id')):
                data['segment_id'] = str(segment_id)
            event_dicts.append(data)

        return event_dicts

    def queue_api_events(
        self,
        events_list: List,
        video_source: str = None,
        segment_id: Union[int, str] = None
    ) -> int:
        """Queue newly-created events for the next wall-clock API batch."""
        if not self.config.enable_api_upload or not events_list:
            return 0

        event_dicts = self._convert_events_to_dicts(
            events_list,
            video_source=video_source,
            segment_id=segment_id
        )
        queued = get_api_batch_uploader().queue_events(event_dicts)
        if queued:
            self.logger.debug(f"Queued {queued} event(s) for the next API boundary")
        return queued

    def export_segment(self, events_list: List, segment_id: Union[int, str],
                       start_dt: datetime.datetime, end_dt: datetime.datetime,
                       source_name: str = None) -> Dict[str, str]:
        """
        Export a segment of events - convenience wrapper for VideoWorker

        Args:
            events_list: List of CountingEvent objects or event dictionaries
            segment_id: Segment identifier
            start_dt: Segment start datetime
            end_dt: Segment end datetime
            source_name: Optional source video filename

        Returns:
            Dictionary of exported file paths
        """
        event_dicts = self._convert_events_to_dicts(events_list)

        # Build the events dictionary in the format expected by export_segment_results
        events_dict = {
            'events': event_dicts,
            '_window_start': start_dt.isoformat() if start_dt else datetime.datetime.now().isoformat(),
            '_window_end': end_dt.isoformat() if end_dt else datetime.datetime.now().isoformat()
        }

        return self.export_segment_results(
            segment_id=segment_id,
            counts={},  # Not used in simplified version
            events=events_dict,
            stats=None,  # Not used in simplified version
            video_source=source_name
        )

    def export_video_summary(self, results: Dict, video_name: str = None) -> Dict[str, str]:
        """
        Export summary for a single video - convenience wrapper for VideoWorker

        Args:
            results: Dictionary with 'final_counts', 'events_summary', 'stats' keys
            video_name: Name of the video file

        Returns:
            Dictionary of exported file paths
        """
        try:
            exported_files = {}
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            prefix = f"{video_name}_" if video_name else ""

            # Create summaries folder
            summaries_folder = self.output_folder / "summaries"
            summaries_folder.mkdir(parents=True, exist_ok=True)

            # Export as JSON
            if 'json' in self.config.export_formats:
                filename = f"{prefix}summary_{timestamp}.json"
                filepath = summaries_folder / filename

                # Build summary data
                summary_data = {
                    'video_name': video_name,
                    'export_timestamp': datetime.datetime.now().isoformat(),
                    'final_counts': results.get('final_counts', {}),
                    'events_summary': results.get('events_summary', {}),
                    'stats': {}
                }

                # Handle ProcessingStats dataclass
                stats = results.get('stats')
                if stats:
                    if hasattr(stats, '__dataclass_fields__'):
                        summary_data['stats'] = asdict(stats)
                    elif hasattr(stats, '__dict__'):
                        summary_data['stats'] = vars(stats)
                    elif isinstance(stats, dict):
                        summary_data['stats'] = stats

                with open(filepath, 'w') as f:
                    json.dump(summary_data, f, indent=2, default=str)

                exported_files['summary_json'] = str(filepath)
                self.logger.info(f"Video summary exported: {filepath}")

            return exported_files

        except Exception as e:
            self.logger.error(f"Failed to export video summary: {e}")
            return {}

    # ========================================================================
    # Segment Export Methods
    # ========================================================================

    def _background_file_export(self, time_str, event_list, window_start, window_end, video_source_str, segment_id_str):
        """Runs the heavy file writing on a background thread so the camera doesn't freeze"""
        try:
            if 'json' in self.config.export_formats:
                self._export_segment_json(time_str, event_list, window_start, window_end,
                                          video_source=video_source_str, segment_id=segment_id_str)
            if 'csv' in self.config.export_formats:
                self._export_segment_csv(time_str, event_list, window_start, window_end,
                                         video_source=video_source_str, segment_id=segment_id_str)
            if 'excel' in self.config.export_formats:
                self._export_segment_excel(time_str, event_list, window_start, window_end,
                                           video_source=video_source_str, segment_id=segment_id_str)
            self.logger.info(f"Background export finished for Segment {segment_id_str}")
        except Exception as e:
            self.logger.error(f"Background file export failed: {e}")

    def export_segment_results(self, segment_id: Union[int, str], counts: Dict, events: Dict, stats: Any,
                              video_source: str = None) -> Dict[str, str]:
        """
        Export event log for a single hourly segment.
        Exports locally first, then uploads to cloud if configured.

        Args:
            segment_id: Segment identifier
            counts: Current counting data (unused in simplified version)
            events: Events data containing the event list
            stats: Processing statistics (unused in simplified version)
            video_source: Optional source video filename

        Returns:
            Dictionary of exported file paths
        """
        try:
            # Extract the event list - it's under 'events' key, not 'event_log'
            event_list = events.get('events', [])

            # Convert segment_id to string for consistency
            segment_id_str = str(segment_id) if segment_id is not None else ''
            video_source_str = video_source or ''

            # Append to master log FIRST
            if self.config.enable_master_log:
                self._append_to_master_log(event_list, video_source=video_source_str, segment_id=segment_id_str)

            # Get window times for filename if available
            window_start = events.get('_window_start', datetime.datetime.now().isoformat())
            window_end = events.get('_window_end', datetime.datetime.now().isoformat())

            # Create time string for filename
            try:
                start_dt = datetime.datetime.fromisoformat(window_start.replace('Z', ''))
                end_dt = datetime.datetime.fromisoformat(window_end.replace('Z', ''))
                time_str = f"{start_dt.strftime('%H%M')}-{end_dt.strftime('%H%M')}_{start_dt.strftime('%Y%m%d')}"
            except:
                # Fallback to current time if parsing fails
                timestamp = datetime.datetime.now()
                time_str = timestamp.strftime("%H%M_%Y%m%d")

            # Export in requested formats with video_source and segment_id
            exported_files = {}

            if 'json' in self.config.export_formats:
                json_path = self._export_segment_json(time_str, event_list, window_start, window_end,
                                                      video_source=video_source_str, segment_id=segment_id_str)
                exported_files['json'] = str(json_path)

            if 'csv' in self.config.export_formats:
                csv_path = self._export_segment_csv(time_str, event_list, window_start, window_end,
                                                    video_source=video_source_str, segment_id=segment_id_str)
                exported_files['csv'] = str(csv_path)

            if 'excel' in self.config.export_formats:
                excel_path = self._export_segment_excel(time_str, event_list, window_start, window_end,
                                                        video_source=video_source_str, segment_id=segment_id_str)
                exported_files['excel'] = str(excel_path)

            self.logger.info(f"Segment {segment_id} event log exported to {len(exported_files)} format(s)")

            # ================================================================
            # ASYNCHRONOUS LOCAL FILE EXPORTS
            # ================================================================
            # We fire this into the ThreadPool so Pandas doesn't freeze the camera!

            safe_event_list = copy.deepcopy(event_list)

            self.file_writer_pool.submit(
                self._background_file_export,
                time_str, safe_event_list, window_start, window_end, video_source_str, segment_id_str
            )

            exported_files = {"status": "Exporting in background"}

            # API events are queued when they are created by the worker. Do not
            # upload this complete segment again here, because doing so would
            # duplicate records already assigned to five-minute API batches.
            if self.config.enable_api_upload:
                exported_files['cloud_status'] = "Handled by wall-clock API batcher"

            return exported_files

        except Exception as e:
            self.logger.error(f"Failed to export segment {segment_id} results: {e}")
            return {}

    def _export_segment_json(self, time_str: str, event_list: List[Dict], window_start: str, window_end: str,
                              video_source: str = '', segment_id: str = '') -> Path:
        """Export event log as JSON with standardized fields"""
        filename = f"events_{time_str}.json"
        filepath = self.segments_folder / filename

        # Standardize all events
        standardized_events = []
        for event in event_list:
            std_event = self._standardize_event_fields(event)
            # Add video_source and segment_id if not present
            if not std_event.get('video_source'):
                std_event['video_source'] = video_source
            if not std_event.get('segment_id'):
                std_event['segment_id'] = segment_id
            standardized_events.append(std_event)

        # Structure the JSON with metadata and standardized events
        data = {
            "event_count": len(standardized_events),
            "events": standardized_events
        }

        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2, default=str)

        return filepath

    def _export_segment_csv(self, time_str: str, event_list: List[Dict], window_start: str, window_end: str,
                            video_source: str = '', segment_id: str = '') -> Path:
        """Export event log as CSV with standardized fields"""
        filename = f"events_{time_str}.csv"
        filepath = self.segments_folder / filename

        # Standard column order
        standard_columns = [
            'actual_datetime', 'event_type', 'track_id', 'class_id', 'class_name',
            'line_name', 'zone_name', 'direction', 'confidence', 'speed', 'speed_units',
            'dwell_seconds', 'video_source', 'segment_id'
        ]

        if not event_list:
            # Write empty CSV with standard headers
            with open(filepath, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(standard_columns)
            return filepath

        # Standardize all events
        standardized_events = []
        for event in event_list:
            std_event = self._standardize_event_fields(event)
            if not std_event.get('video_source'):
                std_event['video_source'] = video_source
            if not std_event.get('segment_id'):
                std_event['segment_id'] = segment_id
            standardized_events.append(std_event)

        # Write to CSV with standard column order
        df = pd.DataFrame(standardized_events)
        df = self._reorder_columns(df)
        df.to_csv(filepath, index=False)

        return filepath

    def _export_segment_excel(self, time_str: str, event_list: List[Dict], window_start: str, window_end: str,
                              video_source: str = '', segment_id: str = '') -> Path:
        """Export event log as Excel file with standardized fields"""
        filename = f"events_{time_str}.xlsx"
        filepath = self.segments_folder / filename

        # Standard column order
        standard_columns = [
            'actual_datetime', 'event_type', 'track_id', 'class_id', 'class_name',
            'line_name', 'zone_name', 'direction', 'confidence', 'speed', 'speed_units',
            'dwell_seconds', 'video_source', 'segment_id'
        ]

        if not event_list:
            # Create empty Excel with standard headers
            df = pd.DataFrame(columns=standard_columns)
            df.to_excel(filepath, index=False, engine='openpyxl')
            return filepath

        # Standardize all events
        standardized_events = []
        for event in event_list:
            std_event = self._standardize_event_fields(event)
            if not std_event.get('video_source'):
                std_event['video_source'] = video_source
            if not std_event.get('segment_id'):
                std_event['segment_id'] = segment_id
            standardized_events.append(std_event)

        # Create DataFrame with standard column order
        df = pd.DataFrame(standardized_events)
        df = self._reorder_columns(df)

        # Write to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Events', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets['Events']
            for idx, col in enumerate(df.columns):
                if idx < 26:  # Excel column limit for single letters
                    max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)

        return filepath

    def export_final_summary(self, results: Dict) -> Dict[str, str]:
        """
        Export final results - combines all event logs from processing

        Args:
            results: Final processing results (supports multiple result structures)

        Returns:
            Dictionary of exported file paths
        """
        try:
            exported_files = {}

            # Create a simple final export with all events combined
            all_events = []

            # Extract events from results - handle multiple possible structures
            if 'video_results' in results:
                # Batch video processing - aggregate from all videos
                for video_result in results.get('video_results', []):
                    if isinstance(video_result, dict):
                        # Get video source name for tagging events
                        video_source = video_result.get('video_path', '')
                        if video_source:
                            video_source = Path(video_source).name
                        
                        events_summary = video_result.get('events_summary', {})
                        video_events = events_summary.get('events', [])
                        
                        # Tag each event with video source if not already present
                        for evt in video_events:
                            if isinstance(evt, dict) and not evt.get('video_source'):
                                evt['video_source'] = video_source
                        
                        all_events.extend(video_events)
            elif 'events_summary' in results:
                # Single video/camera processing
                all_events = results['events_summary'].get('events', [])
            elif 'events' in results and isinstance(results['events'], dict):
                all_events = results['events'].get('events', [])
            elif 'events' in results and isinstance(results['events'], list):
                all_events = results['events']

            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

            # Standardize all events to consistent field format
            standardized_events = []
            for event in all_events:
                standardized_events.append(self._standardize_event_fields(event))

            if 'json' in self.config.export_formats:
                filename = f"all_events_{timestamp}.json"
                filepath = self.output_folder / filename

                data = {
                    "export_timestamp": datetime.datetime.now().isoformat(),
                    "total_events": len(standardized_events),
                    "events": standardized_events
                }

                # Write JSON file
                with open(filepath, 'w') as f:
                    json.dump(data, f, indent=2, default=str)

                exported_files['final_json'] = str(filepath)
                self.logger.info(f"Final JSON exported: {filepath}")

            if 'csv' in self.config.export_formats and standardized_events:
                filename = f"all_events_{timestamp}.csv"
                filepath = self.output_folder / filename

                # Convert to DataFrame with consistent column order
                df = pd.DataFrame(standardized_events)
                df = self._reorder_columns(df)
                df.to_csv(filepath, index=False)

                exported_files['final_csv'] = str(filepath)
                self.logger.info(f"Final CSV exported: {filepath}")

            if 'excel' in self.config.export_formats and standardized_events:
                filename = f"all_events_{timestamp}.xlsx"
                filepath = self.output_folder / filename

                # Convert to DataFrame with consistent column order
                df = pd.DataFrame(standardized_events)
                df = self._reorder_columns(df)
                
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='All_Events', index=False)

                    # Auto-adjust column widths
                    worksheet = writer.sheets['All_Events']
                    for idx, col in enumerate(df.columns):
                        if idx < 26:  # Excel column limit for single letters
                            max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                            worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)

                exported_files['final_excel'] = str(filepath)
                self.logger.info(f"Final Excel exported: {filepath}")

            return exported_files

        except Exception as e:
            self.logger.error(f"Failed to export final summary: {e}")
            return {}

    def _standardize_event_fields(self, event: Dict) -> Dict:
        """
        Standardize event dictionary to have consistent fields in consistent order.
        
        Standard fields: actual_datetime, event_type, track_id, class_id, class_name,
                        line_name, zone_name, direction, confidence, speed, speed_units,
                        dwell_seconds, video_source, segment_id
        """
        # Determine event type
        if event.get('line_name'):
            event_type = 'line_crossing'
        elif event.get('zone_name'):
            event_type = 'zone_entry'
        else:
            event_type = event.get('event_type', 'unknown')
        
        return {
            'actual_datetime': event.get('actual_datetime', ''),
            'event_type': event_type,
            'track_id': event.get('track_id', ''),
            'class_id': event.get('class_id', ''),
            'class_name': event.get('class_name', ''),
            'line_name': event.get('line_name', ''),
            'zone_name': event.get('zone_name', ''),
            'direction': event.get('direction', ''),
            'confidence': event.get('confidence', ''),
            'speed': event.get('speed', 0.0),
            'speed_units': event.get('speed_units', ''),
            'dwell_seconds': self._sanitize_dwell_time(event.get('dwell_seconds', 0.0)),
            'video_source': event.get('video_source', ''),
            'segment_id': event.get('segment_id', '')
        }

    def _reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Reorder DataFrame columns to standard order, adding missing columns as empty.
        
        Standard order: actual_datetime, event_type, track_id, class_id, class_name,
                       line_name, zone_name, direction, confidence, speed, speed_units,
                       dwell_seconds, video_source, segment_id
        """
        standard_columns = [
            'actual_datetime', 'event_type', 'track_id', 'class_id', 'class_name',
            'line_name', 'zone_name', 'direction', 'confidence', 'speed', 'speed_units',
            'dwell_seconds', 'video_source', 'segment_id'
        ]
        
        # Add missing columns with empty values
        for col in standard_columns:
            if col not in df.columns:
                df[col] = ''
        
        # Reorder to standard order, keeping any extra columns at the end
        extra_columns = [col for col in df.columns if col not in standard_columns]
        ordered_columns = standard_columns + extra_columns
        
        return df[ordered_columns]

    def export_live_stats(self, stats_data: Dict) -> str:
        """
        Export current live statistics (kept for compatibility but simplified)
        Just exports the raw stats data without processing
        """
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"live_stats_{timestamp}.json"
        filepath = self.output_folder / filename

        with open(filepath, 'w') as f:
            json.dump(stats_data, f, indent=2, default=str)

        self.logger.info(f"Live stats exported: {filepath}")
        return str(filepath)

    def get_export_summary(self) -> Dict:
        """Get summary of exported files location"""
        summary = {
            "output_folder": str(self.output_folder),
            "segments_folder": str(self.segments_folder),
            "export_formats": self.config.export_formats
        }

        if self.config.enable_master_log:
            summary["master_log"] = str(self.master_log_path)

            # Get event count from master log
            try:
                df = pd.read_excel(self.master_log_path, sheet_name='Events', engine='openpyxl')
                summary["master_log_event_count"] = len(df)
            except:
                summary["master_log_event_count"] = 0

        return summary
